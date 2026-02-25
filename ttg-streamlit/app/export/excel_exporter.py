from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


def generate_full_excel(entries: list) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)

    sections: dict = {}
    for entry in entries:
        sec_name = entry.section.name
        if sec_name not in sections:
            sections[sec_name] = []
        sections[sec_name].append(entry)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    alt_fill = PatternFill(start_color="f5f0ff", end_color="f5f0ff", fill_type="solid")

    for section_name, sec_entries in sections.items():
        ws = wb.create_sheet(title=section_name[:31])
        max_period = max(e.timeslot.period_number for e in sec_entries)

        period_cell = ws.cell(row=1, column=1, value="Period")
        period_cell.font = header_font
        period_cell.fill = header_fill
        period_cell.alignment = center_align
        period_cell.border = thin_border

        for col, day in enumerate(DAYS, start=2):
            cell = ws.cell(row=1, column=col, value=day)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        grid = {(e.timeslot.day_of_week, e.timeslot.period_number): e for e in sec_entries}

        for period in range(1, max_period + 1):
            row = period + 1
            pc = ws.cell(row=row, column=1, value=period)
            pc.font = Font(bold=True)
            pc.alignment = center_align
            pc.border = thin_border
            if period % 2 == 0:
                pc.fill = alt_fill

            for col, day in enumerate(DAYS, start=2):
                entry = grid.get((day, period))
                cell = ws.cell(row=row, column=col)
                if entry:
                    cell.value = f"{entry.subject.code}\n{entry.faculty.name}\n{entry.room.name}"
                cell.alignment = center_align
                cell.border = thin_border
                if period % 2 == 0:
                    cell.fill = alt_fill

        ws.column_dimensions["A"].width = 10
        for col in range(2, 7):
            ws.column_dimensions[get_column_letter(col)].width = 22
        for row in range(1, max_period + 2):
            ws.row_dimensions[row].height = 50

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
